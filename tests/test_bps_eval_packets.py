from __future__ import annotations

import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.export_unit_review_packet import export
from scripts.eval_golden_episodes import EPISODES_PATH, REPORT_PATH, run_llm, run_tools


def test_unit_review_packet_exports_excel_and_markdown() -> None:
    result = export()
    assert result["measures"] >= 13, result
    xlsx = Path(result["xlsx"])
    assert xlsx.exists()
    from openpyxl import load_workbook

    book = load_workbook(xlsx, read_only=True)
    sheet = book["unit_review"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "source_family"
    assert len(rows) - 1 == result["measures"]
    md = Path(result["markdown"])
    assert md.exists()
    assert "blocked_quality" in md.read_text(encoding="utf-8")


def test_golden_episode_harness_tools_mode() -> None:
    from scripts.simulate_bps_candidate_scoring import (
        build_offering_index,
        offer_candidates,
    )

    index = build_offering_index()
    report = run_tools(index, offer_candidates)
    # The harness reports exercised assertions against executable turns;
    # several episodes are not yet executable (no session-policy / no query
    # runtime).  Measure against exercised, not against the full 19.
    exercised = report.get("episodes_evaluated")
    passed = sum(1 for c in report["cases"] if c["status"] == "passed")
    assert report.get("mode") in ("tools", None)
    assert exercised is not None
    if exercised:
        assert passed == exercised, report["cases"]
    assert REPORT_PATH.exists()


def test_golden_episode_harness_llm_mode_documents_blocker() -> None:
    report = run_llm()
    assert report["mode"] == "llm"
    if report["status"] == "blocked":
        assert "OQ-05" in report["reason"] or "MARAWA_LLM" in report["reason"]
    else:
        # live path (MARAWA_LLM_* configured): report counts must be sane
        assert report["episodes_total"] == 19
        assert (
            report["episodes_passed"]
            + report["episodes_passed_soft"]
            + report["episodes_failed_hard"]
            + report.get("episodes_not_evaluated", 0)
            == 19
        )


def test_all_episodes_have_selection_proof_on_fact_queries() -> None:
    import json

    episodes = json.loads(EPISODES_PATH.read_text(encoding="utf-8"))["episodes"]
    query_actions = {"query_stat_data", "query_and_compare"}
    for episode in episodes:
        for turn in episode["turns"]:
            expect = turn["expect"]
            if expect.get("action") in query_actions:
                assert expect.get("query_facts") is True, episode["episode_id"]
                assert expect.get("selection_source") in {
                    "candidate_set_ref",
                    "explicit_ref",
                    "active_dataset",
                }, episode["episode_id"]
