from __future__ import annotations

from openpyxl import load_workbook

from scripts.export_bps_catalog_excel import SHEETS, write_workbook


def test_catalog_workbook_contains_metadata_sheets_not_fact_data(tmp_path) -> None:
    output = tmp_path / "catalog.xlsx"
    catalogs = {
        "SIMDASI": [["3.1.1", "Jumlah Penduduk", "Penduduk", "Sosial", "2024, 2025", "2026-01-01"]],
        "DYNAMIC": [["29", "Jumlah Penduduk", "Kependudukan", "jiwa"]],
        "CENSUS": [["sp2010", "Sensus Penduduk 2010", "1", "Topik", "10", "Dataset"]],
        "PUBLICATION": [["pub-1", "Publikasi Contoh", "2026-01-01", "2026-01-02"]],
    }

    write_workbook(catalogs, output, generated_at="2026-08-15T00:00:00Z")

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["PETUNJUK", *SHEETS]
    assert workbook["SIMDASI"].max_row == 2
    assert workbook["SIMDASI"]["A2"].value == "3.1.1"
    assert workbook["DYNAMIC"]["A2"].value == "29"
    assert workbook["DYNAMIC"].max_column == 4
    assert workbook["CENSUS"]["E2"].value == "10"
    assert workbook["PUBLICATION"]["A2"].value == "pub-1"
    all_headers = [cell.value for sheet in SHEETS for cell in workbook[sheet][1]]
    assert not any("definisi" in str(header).lower() or "nilai" in str(header).lower() for header in all_headers)
    assert workbook["SIMDASI"].freeze_panes == "A2"
    assert workbook["SIMDASI"].auto_filter.ref == "A1:F2"


def test_catalog_workbook_keeps_empty_catalogue_as_header_only(tmp_path) -> None:
    output = tmp_path / "empty.xlsx"

    write_workbook({sheet: [] for sheet in SHEETS}, output, generated_at="2026-08-15T00:00:00Z")

    workbook = load_workbook(output, data_only=True)
    for sheet in SHEETS:
        assert workbook[sheet].max_row == 1
        assert workbook[sheet].auto_filter.ref is None
