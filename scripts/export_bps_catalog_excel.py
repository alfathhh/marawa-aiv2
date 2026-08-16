#!/usr/bin/env python3
"""Export BPS catalogue identifiers to an Excel workbook; never exports facts/cells."""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from workers.ingestion.bps_storage import load_postgres_dsn

POSTGRES_ENV = Path("/home/ubuntu/.config/marawa-ai/postgres.env")
OUTPUT = ROOT / "data" / "reports" / "BPS_CATALOG_TABEL_1306.xlsx"
SHEETS = ("SIMDASI", "DYNAMIC", "CENSUS", "PUBLICATION")
HEADERS: dict[str, list[str]] = {
    "SIMDASI": ["Kode Tabel", "Judul Tabel", "Bab", "Subjek", "Tahun Tersedia", "Update Terakhir BPS"],
    "DYNAMIC": ["ID Indikator", "Nama Indikator", "Subjek", "Unit"],
    "CENSUS": ["ID Event", "Nama Event", "ID Topik", "Nama Topik", "ID Dataset", "Nama Dataset"],
    "PUBLICATION": ["ID Publikasi", "Judul Publikasi", "Tanggal Rilis", "Tanggal Update"],
}
WIDTHS: dict[str, list[int]] = {
    "SIMDASI": [16, 66, 28, 34, 22, 23],
    "DYNAMIC": [16, 72, 34, 20],
    "CENSUS": [16, 32, 14, 44, 16, 78],
    "PUBLICATION": [30, 72, 18, 18],
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")


def safe_cell(value: Any) -> Any:
    """Avoid formula interpretation when an upstream label begins with Excel syntax."""
    if not isinstance(value, str):
        return value
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def _years(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    years: set[int] = set()
    for item in value:
        try:
            years.add(int(float(str(item))))
        except (TypeError, ValueError):
            continue
    return ", ".join(str(year) for year in sorted(years))


def collect_catalogs() -> dict[str, list[list[Any]]]:
    """Read only catalogue metadata from local normalized tables, never facts."""
    with psycopg.connect(load_postgres_dsn(POSTGRES_ENV)) as connection:
        simdasi = [
            [row[0], row[1], row[2], row[3], _years(row[4]), row[5]]
            for row in connection.execute(
                """
                SELECT table_code, title, chapter, subject, available_years, raw->>'latest_update'
                FROM bps_simdasi_tables
                WHERE region_code=%s
                ORDER BY string_to_array(table_code, '.')::int[]
                """,
                ("1306000",),
            ).fetchall()
        ]
        dynamic = [
            [row[0], row[1], row[2], row[3]]
            for row in connection.execute(
                """
                SELECT variable_id, title, subject_name, coalesce(unit_canonical, unit)
                FROM bps_dynamic_variables
                WHERE domain=%s
                ORDER BY variable_id::bigint
                """,
                ("1306",),
            ).fetchall()
        ]
        census = [
            [row[0], row[1], row[2], row[3], row[4], row[5]]
            for row in connection.execute(
                """
                SELECT e.event_id, e.event_name, t.topic_id, t.topic_name, d.dataset_id, d.dataset_name
                FROM bps_census_datasets d
                JOIN bps_census_topics t ON t.event_id=d.event_id AND t.topic_id=d.topic_id
                JOIN bps_census_events e ON e.event_id=d.event_id
                ORDER BY e.event_id, t.topic_id, d.dataset_id
                """
            ).fetchall()
        ]
        publication = [
            [row[0], row[1], row[2], row[3]]
            for row in connection.execute(
                """
                SELECT publication_id, title, release_date, updated_date
                FROM bps_publications
                WHERE domain=%s
                ORDER BY release_date DESC NULLS LAST, publication_id
                """,
                ("1306",),
            ).fetchall()
        ]
    return {"SIMDASI": simdasi, "DYNAMIC": dynamic, "CENSUS": census, "PUBLICATION": publication}


def add_instructions(workbook: Workbook, generated_at: str) -> None:
    sheet = workbook.active
    sheet.title = "PETUNJUK"
    rows = [
        ["Katalog BPS Padang Pariaman — Metadata Saja"],
        ["Dibuat", generated_at],
        [],
        ["File ini tidak memuat nilai/facts/cell data. Gunakan identifier pada kolom pertama saat melaporkan update."],
        [],
        ["Format laporan update ke Lord"],
        ["SIMDASI 3.1.1 tahun 2026"],
        ["DYNAMIC 29 tahun 2026"],
        ["CENSUS sp2010 dataset 10"],
        ["PUBLICATION 9d824be2b30029991c8aed8a"],
        [],
        ["Jika yang direvisi adalah tahun lama, selalu sebutkan tahunnya agar hanya resource exact itu yang ditarik."],
        ["Tidak ada BPS cronjob aktif; update hanya berjalan setelah Tah memberi instruksi manual."],
    ]
    for row in rows:
        sheet.append(row)
    sheet.merge_cells("A1:D1")
    sheet["A1"].font = Font(bold=True, size=14, color="1F1F1F")
    sheet["A1"].fill = TITLE_FILL
    sheet["A6"].font = Font(bold=True, color="FFFFFF")
    sheet["A6"].fill = HEADER_FILL
    sheet.column_dimensions["A"].width = 105
    sheet.column_dimensions["B"].width = 30
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(catalogs: dict[str, list[list[Any]]], output: Path, *, generated_at: str) -> None:
    workbook = Workbook()
    add_instructions(workbook, generated_at)
    for sheet_name in SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        headers = HEADERS[sheet_name]
        sheet.append(headers)
        for raw_row in catalogs.get(sheet_name, []):
            row = list(raw_row)[:len(headers)]
            row.extend([None] * (len(headers) - len(row)))
            sheet.append([safe_cell(value) for value in row])
        for index, width in enumerate(WIDTHS[sheet_name], 1):
            sheet.column_dimensions[chr(64 + index)].width = width
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 34
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet.max_row > 1:
            end_col = chr(64 + len(headers))
            ref = f"A1:{end_col}{sheet.max_row}"
            sheet.auto_filter.ref = ref
            table = Table(displayName=f"Catalog{sheet_name.title()}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.properties.creator = "MARAWA AI"
    workbook.properties.title = "Katalog Metadata BPS Padang Pariaman"
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export BPS catalogue identifiers only (no table facts).")
    parser.add_argument("--output", type=Path, default=OUTPUT)
    args = parser.parse_args()
    catalogs = collect_catalogs()
    write_workbook(catalogs, args.output, generated_at=datetime.now(timezone.utc).isoformat())
    summary = {sheet: len(rows) for sheet, rows in catalogs.items()}
    print({"output": str(args.output), "metadata_rows": summary, "facts_exported": 0})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
